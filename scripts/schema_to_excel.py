"""
schema_to_excel.py

Exports the merged CoreMeta4Cat LinkML schema to an Excel workbook.

Sheet order:
  1. Introduction  -- overview, how to use the workbook
  2. Legend        -- colour coding and column explanations
  3. CoreMeta4Cat  -- CatalysisDataset top-level fields (the minimal cross-cutting layer)
  4. Synthesis
  5. Characterization
  6. Reaction
  7. Simulation

Columns (sheets 3-7):
  label                      human-readable slot or class name
  parent                     parent slot/class (empty for top-level entries)
  mandatory, recommended, optional  M / R / O classification
  range                      LinkML range type
  slot_uri / class_uri       CURIE for the term
  description                documentation string

Output:
  docs/assets/coremeta4cat_vocabulary.xlsx  (overwrites the file in place)

Run:
    just schema-to-excel
  or directly:
    uv run python scripts/schema_to_excel.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).parent
_ROOT = _HERE.parent
sys.path.insert(0, str(_HERE))

from generate_schema_docs import (  # noqa: E402
    get_all_class_slots,
    get_class_ranged_slot_usage,
    get_slot_details,
    get_subclasses,
    is_mixin,
    load_merged_schema,
    snake_to_readable,
)

# ─────────────────────────────────────────────────────────────────────────────
# Styles
# ─────────────────────────────────────────────────────────────────────────────

FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER   = Font(color="FFFFFF", bold=True)
FILL_TITLE    = PatternFill("solid", fgColor="2E75B6")
FONT_TITLE    = Font(color="FFFFFF", bold=True, size=14)
FILL_SECTION  = PatternFill("solid", fgColor="BDD7EE")
FONT_SECTION  = Font(bold=True)
FILL_M        = PatternFill("solid", fgColor="FCE4D6")
FILL_R        = PatternFill("solid", fgColor="DEEAF1")
FILL_O        = PatternFill("solid", fgColor="E2EFDA")
FILL_M_LEGEND = PatternFill("solid", fgColor="C00000")
FILL_R_LEGEND = PatternFill("solid", fgColor="2E75B6")
FILL_O_LEGEND = PatternFill("solid", fgColor="70AD47")
FONT_WHITE    = Font(color="FFFFFF", bold=True)

HEADERS = [
    "label",
    "parent",
    "mandatory, recommended, optional",
    "range",
    "slot_uri / class_uri",
    "description",
]

# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _slot_mro(schema: dict, class_name: str, slot_name: str) -> str:
    class_def = schema.get("classes", {}).get(class_name, {})
    usage = (class_def.get("slot_usage") or {}).get(slot_name) or {}
    slot_def = get_slot_details(schema, slot_name)
    required    = usage.get("required",    slot_def.get("required",    False))
    recommended = usage.get("recommended", slot_def.get("recommended", False))
    if required:
        return "M"
    if recommended:
        return "R"
    return "O"


def _write_data_headers(ws) -> None:
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _append_row(ws, row_data: list, mro: str) -> None:
    ws.append(row_data)
    fill = {"M": FILL_M, "R": FILL_R, "O": FILL_O}.get(mro)
    if fill:
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────────────────────────────────────
# Introduction sheet
# ─────────────────────────────────────────────────────────────────────────────

def build_intro_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(title="Introduction")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    def _title(row: int, text: str) -> None:
        ws.row_dimensions[row].height = 28
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_TITLE
        c.fill = FILL_TITLE
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def _section(row: int, text: str) -> None:
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_SECTION
        c.fill = FILL_SECTION
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def _row(row: int, label: str, value: str) -> None:
        ws.row_dimensions[row].height = 40
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        vc = ws.cell(row=row, column=2, value=value)
        vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _blank(row: int) -> None:
        ws.row_dimensions[row].height = 8

    r = 1
    _title(r, "CoreMeta4Cat Vocabulary Reference Workbook"); r += 1
    _blank(r); r += 1

    _section(r, "About this workbook"); r += 1
    _row(r, "Project", "CoreMeta4Cat is a community-driven metadata initiative under NFDI4Cat that defines the minimum information required for reporting catalysis research data. It is built on the FAIR principles (Findable, Accessible, Interoperable, Reusable)."); r += 1
    _row(r, "Purpose", "This workbook is a structured reference overview of the CoreMeta4Cat vocabulary hierarchy, organised by data class. It is NOT a data entry form. Use it as a lookup reference when designing or annotating your own data sheets."); r += 1
    _row(r, "Ground truth", "The LinkML schema files in src/coremeta4cat/schema/ are the authoritative source. This workbook is generated automatically from the schema via: just schema-to-excel"); r += 1
    _row(r, "Source", "https://github.com/nfdi4cat/CoreMeta4Cat"); r += 1
    _blank(r); r += 1

    _section(r, "Sheet overview"); r += 1
    _row(r, "Introduction", "This sheet — project overview and guidance."); r += 1
    _row(r, "Legend", "Explanation of colour coding and column meanings."); r += 1
    _row(r, "CoreMeta4Cat", "The minimal cross-cutting fields that apply to every catalysis dataset, regardless of data class (based on CatalysisDataset)."); r += 1
    _row(r, "Synthesis", "Metadata fields for catalyst preparation / synthesis experiments."); r += 1
    _row(r, "Characterization", "Metadata fields for analytical characterization of catalysts."); r += 1
    _row(r, "Reaction", "Metadata fields for catalytic reaction testing and performance evaluation."); r += 1
    _row(r, "Simulation", "Metadata fields for computational / theoretical catalysis studies."); r += 1
    _blank(r); r += 1

    _section(r, "How to read the data sheets"); r += 1
    _row(r, "label", "Human-readable name of the metadata field or class."); r += 1
    _row(r, "parent", "The parent field this entry belongs to. Empty means it is a top-level field."); r += 1
    _row(r, "M / R / O", "Whether the field is Mandatory (must be reported), Recommended (strongly encouraged), or Optional. See the Legend sheet for colour coding."); r += 1
    _row(r, "range", "The expected value type — a primitive (string, float, integer) or a sub-class name that expands into its own set of fields."); r += 1
    _row(r, "slot_uri / class_uri", "The CURIE (Compact URI) linking this field to a term in an established ontology (e.g. Voc4Cat, CHMO, OBI)."); r += 1
    _row(r, "description", "Documentation string from the schema explaining the purpose of the field."); r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Legend sheet
# ─────────────────────────────────────────────────────────────────────────────

def build_legend_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(title="Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 70

    def _title(row: int, text: str) -> None:
        ws.row_dimensions[row].height = 28
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_TITLE
        c.fill = FILL_TITLE
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    def _section(row: int, text: str) -> None:
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_SECTION
        c.fill = FILL_SECTION
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    def _color_row(row: int, label: str, fill: PatternFill, font: Font, example: str, desc: str) -> None:
        ws.row_dimensions[row].height = 36
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill
        lc.font = font
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ec = ws.cell(row=row, column=2, value=example)
        ec.fill = fill
        ec.font = font
        ec.alignment = Alignment(horizontal="center", vertical="center")
        dc = ws.cell(row=row, column=3, value=desc)
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _col_row(row: int, col_name: str, desc: str) -> None:
        ws.row_dimensions[row].height = 40
        lc = ws.cell(row=row, column=1, value=col_name)
        lc.font = Font(bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        dc = ws.cell(row=row, column=3, value=desc)
        dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _blank(row: int) -> None:
        ws.row_dimensions[row].height = 8

    r = 1
    _title(r, "Legend"); r += 1
    _blank(r); r += 1

    _section(r, "Colour coding — M / R / O"); r += 1
    _color_row(r, "Mandatory (M)", FILL_M_LEGEND, FONT_WHITE,
               "M", "Must be reported. The field is required to produce a valid, schema-compliant record."); r += 1
    _color_row(r, "Recommended (R)", FILL_R_LEGEND, FONT_WHITE,
               "R", "Strongly encouraged. Omitting these fields significantly reduces the findability and reusability of the data."); r += 1
    _color_row(r, "Optional (O)", FILL_O_LEGEND, FONT_WHITE,
               "O", "Useful additional context. Provide if available; not required for a valid record."); r += 1
    _blank(r); r += 1

    _section(r, "Column descriptions"); r += 1
    _col_row(r, "label",
             "Human-readable name of the metadata field or class, derived from the LinkML slot or class name."); r += 1
    _col_row(r, "parent",
             "The parent field this entry belongs to in the hierarchy. An empty parent means the field is at the top level of the data class."); r += 1
    _col_row(r, "mandatory, recommended, optional",
             "M = Mandatory, R = Recommended, O = Optional. Derived from the 'required' and 'recommended' attributes in the LinkML schema."); r += 1
    _col_row(r, "range",
             "The expected value type. Primitives (string, float, integer, boolean, uri) are leaf values. A class name means the field expands into a structured sub-record defined by the named class."); r += 1
    _col_row(r, "slot_uri / class_uri",
             "CURIE (Compact URI) linking this field to a term in an established ontology. Click the linked term in the schema documentation for the full URI."); r += 1
    _col_row(r, "description",
             "Documentation string from the schema explaining the meaning and purpose of the field."); r += 1
    _blank(r); r += 1

    _section(r, "Notes"); r += 1
    ws.row_dimensions[r].height = 32
    nc = ws.cell(row=r, column=1,
                 value="This workbook is generated automatically from the LinkML schema. The schema is the authoritative source — do not edit this file manually. To regenerate: just schema-to-excel")
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)


# ─────────────────────────────────────────────────────────────────────────────
# CoreMeta4Cat (CatalysisDataset) sheet
# ─────────────────────────────────────────────────────────────────────────────

def build_catcore_sheet(wb: openpyxl.Workbook, schema: dict) -> None:
    """
    The minimal cross-cutting layer — fields on CatalysisDataset that apply
    to every catalysis dataset regardless of data class.
    """
    ws = wb.create_sheet(title="CoreMeta4Cat")
    _write_data_headers(ws)

    classes  = schema.get("classes", {})
    cat_def  = classes.get("CatalysisDataset", {})
    slot_usage = cat_def.get("slot_usage") or {}

    rows: list = []
    for su_name, su_def in slot_usage.items():
        if not su_def:
            continue
        required    = su_def.get("required",    False)
        recommended = su_def.get("recommended", False)
        mro = "M" if required else ("R" if recommended else "O")
        rng  = su_def.get("range", "")
        uri  = su_def.get("slot_uri", "") or (classes.get(rng, {}).get("class_uri", "") if rng else "")
        desc = su_def.get("description", "")
        rows.append((snake_to_readable(su_name), "", mro, rng, uri, desc))

    seen: set = set()
    for row in rows:
        key = (row[0], row[1])
        if key in seen:
            continue
        seen.add(key)
        _append_row(ws, list(row), row[2])

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Data-class sheet builder (Synthesis, Characterization, Reaction, Simulation)
# ─────────────────────────────────────────────────────────────────────────────

def _collect_rows(
    schema: dict,
    class_name: str,
    parent_label: str,
    context_class: str,
    seen: set,
    rows: list,
    depth: int = 0,
) -> None:
    if class_name in seen or depth > 8:
        return
    seen = seen | {class_name}
    classes = schema.get("classes", {})

    direct_slots = get_all_class_slots(schema, class_name)
    for slot_name in direct_slots:
        mro = _slot_mro(schema, context_class if depth == 0 else class_name, slot_name)
        slot_def = get_slot_details(schema, slot_name)
        range_type = slot_def.get("range", "string")
        uri  = slot_def.get("slot_uri", "")
        desc = slot_def.get("description", "")
        label = snake_to_readable(slot_name)
        rows.append((label, parent_label, mro, range_type, uri, desc))

        if range_type and range_type in classes and not is_mixin(schema, range_type):
            class_def   = classes[range_type]
            class_label = snake_to_readable(range_type)
            class_uri   = class_def.get("class_uri", "")
            class_desc  = class_def.get("description", "")
            rows.append((class_label, label, mro, "", class_uri, class_desc))
            _collect_rows(schema, range_type, class_label, context_class, seen, rows, depth + 1)
            for sub in get_subclasses(schema, range_type):
                sub_def   = classes.get(sub, {})
                sub_label = snake_to_readable(sub)
                sub_uri   = sub_def.get("class_uri", "")
                sub_desc  = sub_def.get("description", "")
                rows.append((sub_label, class_label, mro, "", sub_uri, sub_desc))
                _collect_rows(schema, sub, sub_label, context_class, seen | {range_type}, rows, depth + 2)

    for su_name, synthetic in get_class_ranged_slot_usage(schema, class_name):
        if su_name in direct_slots:
            continue
        rng  = synthetic.get("range", "")
        mro  = _slot_mro(schema, class_name, su_name)
        uri  = synthetic.get("slot_uri", "")
        desc = synthetic.get("description", "")
        su_label = snake_to_readable(su_name)
        rows.append((su_label, parent_label, mro, rng, uri, desc))
        if rng and rng in classes and not is_mixin(schema, rng):
            class_def   = classes[rng]
            class_label = snake_to_readable(rng)
            class_uri   = class_def.get("class_uri", "")
            class_desc  = class_def.get("description", "")
            rows.append((class_label, su_label, mro, "", class_uri, class_desc))
            _collect_rows(schema, rng, class_label, class_name, seen, rows, depth + 1)
            for sub in get_subclasses(schema, rng):
                sub_def   = classes.get(sub, {})
                sub_label = snake_to_readable(sub)
                sub_uri   = sub_def.get("class_uri", "")
                sub_desc  = sub_def.get("description", "")
                rows.append((sub_label, class_label, mro, "", sub_uri, sub_desc))
                _collect_rows(schema, sub, sub_label, class_name, seen | {rng}, rows, depth + 2)


def build_sheet(wb: openpyxl.Workbook, schema: dict, class_name: str) -> None:
    ws = wb.create_sheet(title=class_name)
    _write_data_headers(ws)

    rows: list = []
    _collect_rows(schema, class_name, "", class_name, set(), rows)

    seen_labels: set = set()
    for row in rows:
        key = (row[0], row[1])
        if key in seen_labels:
            continue
        seen_labels.add(key)
        _append_row(ws, list(row), row[2])

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main(schema_dir: str, output_path: str) -> None:
    print(f"\nLoading CoreMeta4Cat modules from: {schema_dir}")
    schema = load_merged_schema(schema_dir)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  Building sheet: Introduction")
    build_intro_sheet(wb)

    print("  Building sheet: Legend")
    build_legend_sheet(wb)

    print("  Building sheet: CoreMeta4Cat")
    build_catcore_sheet(wb, schema)

    for cls in ["Synthesis", "Characterization", "Reaction", "Simulation"]:
        print(f"  Building sheet: {cls}")
        build_sheet(wb, schema, cls)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    schema_dir  = str(_ROOT / "src" / "coremeta4cat" / "schema")
    output_path = str(_ROOT / "docs" / "assets" / "coremeta4cat_vocabulary.xlsx")
    main(schema_dir, output_path)
